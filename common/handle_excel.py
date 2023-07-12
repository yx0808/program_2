# 该模块用于读取excel

import openpyxl

class HandleExcel:
    def __init__(self,filename,sheetname):
        self.filename=filename
        self.sheetname=sheetname

    def read_data(self):
        workbook=openpyxl.load_workbook(self.filename)
        sh=workbook[self.sheetname]
        res=list(sh.rows)
        title=[i.value for i in res[0]]

        ca=[]
        for item in res[1:]:
            data=[i.value for i in item]
            dic=dict(zip(title,data))
            ca.append(dic)
        return ca

    def write_data(self,row,column,value):
      workbook=openpyxl.load_workbook(self.filename)
      sh=workbook[self.sheetname]
      sh.cell(row=row,column=column,value=value)
      workbook.save(self.filename)